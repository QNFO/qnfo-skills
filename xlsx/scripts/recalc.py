#!/usr/bin/env python3
"""Recalculate Excel formulas using LibreOffice and scan for errors.

Usage: python recalc.py <excel_file> [timeout_seconds]
Returns JSON with status, formula count, and error details.
Works on Windows, Linux, and macOS.
"""

import json
import os
import subprocess
import sys
import platform
from pathlib import Path

TIMEOUT = 60 if len(sys.argv) < 3 else int(sys.argv[2])
FILE_PATH = sys.argv[1] if len(sys.argv) > 1 else None

if not FILE_PATH:
    print(json.dumps({"status": "error", "message": "Usage: recalc.py <excel_file> [timeout_seconds]"}))
    sys.exit(1)

FILE_PATH = str(Path(FILE_PATH).resolve())
if not os.path.exists(FILE_PATH):
    print(json.dumps({"status": "error", "message": f"File not found: {FILE_PATH}"}))
    sys.exit(1)

# Locate LibreOffice - platform-aware paths
if platform.system() == "Windows":
    LO_PATHS = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
elif platform.system() == "Darwin":
    LO_PATHS = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
else:
    LO_PATHS = [
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/snap/bin/libreoffice",
    ]

LO_BIN = None
for p in LO_PATHS:
    if os.path.exists(p):
        LO_BIN = p
        break
if LO_BIN is None:
    # Try PATH lookup as fallback
    try:
        if platform.system() == "Windows":
            result = subprocess.run(["where", "soffice.exe"], capture_output=True, text=True)
        else:
            result = subprocess.run(["which", "soffice"], capture_output=True, text=True)
        if result.returncode == 0 and result.stdout.strip():
            LO_BIN = result.stdout.strip().split("\n")[0]
    except Exception:
        pass

if not LO_BIN:
    print(json.dumps({
        "status": "error",
        "message": "LibreOffice not found. Install from https://www.libreoffice.org/download/ and re-run."
    }))
    sys.exit(1)

# Phase 1: Scan formulas using openpyxl (available on all platforms, no LO needed)
try:
    import openpyxl
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)

total_formulas = 0
error_cells = {}

try:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1
                # Check for error strings that may have been written
                err_str = str(cell.value).strip().upper()
                if err_str in ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"):
                    loc = f"{sheet_name}!{cell.coordinate}"
                    err_type = err_str if not err_str.endswith("!") else err_str + "!"
                    if err_type not in error_cells:
                        error_cells[err_type] = {"count": 0, "locations": []}
                    error_cells[err_type]["count"] += 1
                    error_cells[err_type]["locations"].append(loc)

    wb.close()

    # Phase 2: Recalculate with LibreOffice
    import tempfile

    with tempfile.TemporaryDirectory() as tmpdir:
        # Copy file to temp location to avoid locking
        import shutil

        tmp_file = os.path.join(tmpdir, os.path.basename(FILE_PATH))
        shutil.copy2(FILE_PATH, tmp_file)

        cmd = [
            LO_BIN,
            "--headless",
            "--norestore",
            "--calc",
            "--convert-to", "xlsx",
            "--outdir", tmpdir,
            tmp_file,
        ]

        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=TIMEOUT)

            # Check for recalculated file
            converted = os.path.join(tmpdir, os.path.basename(FILE_PATH))
            if os.path.exists(converted) and converted != FILE_PATH:
                shutil.copy2(converted, FILE_PATH)

            # Re-scan for errors after recalculation
            wb2 = openpyxl.load_workbook(FILE_PATH, data_only=True)
            post_errors = {}
            for sheet_name in wb2.sheetnames:
                ws2 = wb2[sheet_name]
                for row in ws2.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        err_str = str(cell.value).strip().upper()
                        if err_str in ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"):
                            loc = f"{sheet_name}!{cell.coordinate}"
                            err_type = err_str if not err_str.endswith("!") else err_str + "!"
                            if err_type not in post_errors:
                                post_errors[err_type] = {"count": 0, "locations": []}
                            post_errors[err_type]["count"] += 1
                            post_errors[err_type]["locations"].append(loc)
            wb2.close()

            total_errors = sum(e["count"] for e in post_errors.values())

            result = {
                "status": "errors_found" if total_errors > 0 else "success",
                "total_formulas": total_formulas,
                "total_errors": total_errors,
                "error_summary": post_errors if post_errors else {},
                "recalculated": True,
            }

            if not post_errors and error_cells:
                result["pre_recalc_errors"] = sum(e["count"] for e in error_cells.values())
                result["note"] = "Pre-existing errors resolved by recalculation."

            print(json.dumps(result, indent=2))

        except subprocess.TimeoutExpired:
            print(json.dumps({
                "status": "timeout",
                "total_formulas": total_formulas,
                "total_errors": 0,
                "error_summary": error_cells if error_cells else {},
                "note": f"LibreOffice recalculation timed out after {TIMEOUT}s. Errors shown are from pre-recalc scan. Open file in Calc and press Ctrl+Shift+F9 for full recalc.",
            }))

except Exception as e:
    print(json.dumps({"status": "error", "message": str(e)}))
    sys.exit(1)
